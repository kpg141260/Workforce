# wfm_xl.py
# Copyright 2020 by Peter Gossler. All rights reserved.

import	logging
import	os
from	openpyxl import load_workbook, Workbook
from	openpyxl.styles import PatternFill
from	datetime import datetime, timedelta
import	wfm_helpers
#import threading

class wfm_xl:
#  ================== Initialise class ==================
	def __init__ (self, env:dict, cnf:dict, logger, xlfile=None):
		try:
			self.__lgr		= logger
			self.__env		= env.copy()
			self.__cnf		= cnf.copy()
			self.__dic_fc	= {}
			self.__wb		= Workbook()
			self.__ws		= Workbook.worksheets
			self.__wb_fname	= ''
			self.__date_ary	= []
			self.__time_ary	= []
			self.__cnt_iv	= 0
			self.__max_days	= 0
			self.__max_rows	= 0
			self.__max_row	= 0
			self.__min_row	= 0
			self.__max_cols	= 0
			self.__max_col	= 0
			self.__min_col	= 0
			self.__max_sht	= 0
			self.__svc_time	= 0
			self.__svc_intv	= 0
			self.__row_offs	= 0
			self.__strt_tm	= datetime.now()
			self.__strt_dt	= datetime.now()
			self.__end_tm	= datetime.now()
			self.__isReady	= False
			self.__hasDict	= False
			self.__hasFW	= False
			self.__hasFW	= False
			self.__isSource	= False
			self.__isSimple	= env['excel']['Simple']
			if xlfile is not None:
				self.__wb_fname = xlfile
			else:
				self.__wb_fname = cnf['f_xl_full']
			# Pre-process the excel data source - get max rows, max columns, start time and Service Interval
			if self.__env['excel']['Simple']:
				self.__lgr.info (self.__cnf['res_strings']['excel']['0013'])
				self.xlPreProcessSimple ()
			else:
				self.__lgr.info (self.__cnf['res_strings']['excel']['0014'])
				self.xlPreProcessComplex ()
			# preprocess = threading.Thread(target=self.xlPreProcessSource, daemon=True)
		except Exception as e:
			self.__isReady = False
			raise Exception (e)
			
# ================== General Methods ==================
	def __del__ (self):
		if (self.__isReady):
			self.__wb.close()
			del self.__ws
			del self.__wb
			del self.__wb_fname
			del self.__date_ary
			del self.__strt_tm
			del self.__strt_dt
			del self.__cnf
			del self.__dic_fc
			del self.__env
			del self.__lgr
	
	def __str__ (self):
		self.__wb.worksheets.count
		return f"Class: wfm_xl, Current sheet: {self.__ws.sheet_name}, Total sheets: {self.__wb.worksheets.count}"

	def max_days (self):
		return self.__max_days

	def interval_count (self):
		return self.__cnt_iv

	def max_rows (self):
		return self.__max_rows

	def max_row (self):
		return self.__max_row

	def max_col (self):
		return self.__max_col

	def max_columns (self):
		return self.__max_cols

	def ServiceInterval (self):
		return self.__svc_intv
	
	def ServiceTime (self):
		return self.__svc_time
	
	def getWorksheet (self):
		return self.__ws

	def getWorkbook (self):
		return self.__wb

	def getFcDictionary (self):
		return self.__dic_fc

	def getStartDate (self):
		return self.__strt_dt
	
	def setStartDate (self, date:datetime):
		self.__strt_dt = date

	def getDateArray (self):
		return self.__date_ary

	def hasFramework (self):
		return self.__hasFW

	def hasCallData (self):
		return self.__hasDict

# ================== Pre-process Excel Source Data Simple -- NOT THREAD SAFE ==================
	def xlPreProcessSimple (self):
		try:
			# load the source forecast data from xl
			self.__lgr.info (str.format(self.__cnf['res_strings']['info']['0038'], self.__wb_fname))
			self.__wb = load_workbook(self.__wb_fname, data_only=True)
			self.__lgr.debug (str.format(self.__cnf['res_strings']['excel']['0003'], self.__wb_fname))

			# Check if there is a framework sheet provided in teh Excel workbook and if we should use it
			if self.__env['excel']['useFramework'] and not self.__hasFW:
				# Check if there is a Framework Sheet
				sheets = self.__wb.get_sheet_names()
				for sheet in sheets:
					if sheet == self.__env['excel']['FWS_name']:
						self.__hasFW = True
						self.__lgr.info (str.format(self.__cnf['res_strings']['excel']['0001'], self.__env['excel']['FWS_name'], self.__wb_fname))
				# Yes, there is - let's use it
				if self.__hasFW:
					self.xlReadFrameworkData()

			# Select the active data sheet
			self.__ws = self.__wb[self.__env['excel']['SD_name']]
		except:
			# Something went wrong
			err = str.format(self.__cnf['res_strings']['errors']['0017'], self.__wb_fname)
			self.__lgr.error (err)
			raise Exception (err)
		# Determine the row and column count
		self.__max_rows = self.__ws.max_row
		self.__max_cols = self.__ws.max_column
		# Determine the Service Interval from xl sheet
		try:
			# Check that all columns contain the valid format and are not blank
			for c in range (2, self.__max_cols, 1):
				# Seems to be more logical for log-file to check for empty cells first, so do not change the order of the two if statements
				if (self.__ws.cell(1, c).value is None):
					self.__ws.cell(1, c).fill = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')
					self.__lgr.warning (str.format(self.__cnf['res_strings']['warnings']['0005'], self.__ws.cell(1, c).column_letter, self.__ws.cell(1, c).row, c - 2))
					break
				if (not self.__ws.cell(1, c).is_date):
					self.__ws.cell(1, c).fill = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')
					self.__lgr.warning (str.format(self.__cnf['res_strings']['warnings']['0006'], self.__ws.cell(1, c).column_letter, self.__ws.cell(1, c).row, type(self.__ws.cell(1, c).value), c - 2))
					break
			# Get the start date of the Excel source data
			self.__strt_dt = self.__ws.cell(1, 2).value
			# assign correct max column value, if any
			self.__max_cols = c - 2 
			self.__max_col  = c
			# Check that the first row contains time data
			c = self.__ws.min_column
			for r in range (2, self.__max_rows, 1):
				if (self.__ws.cell(row=r, column=c).value is None):
					self.__ws.cell(row=r, column=c).fill = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')
					self.__lgr.warning (str.format(self.__cnf['res_strings']['warnings']['0007'], self.__ws.cell(row=r, column=c).column_letter, self.__ws.cell(row=r, column=c).row, r - 2))
					break           
				if (not self.__ws.cell(row=r, column=c).is_date):
					self.__ws.cell(1, c).fill = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')
					self.__lgr.warning (str.format(self.__cnf['res_strings']['warnings']['0008'], self.__ws.cell(row=r, column=c).column_letter, self.__ws.cell(row=r, column=c).row, type(self.__ws.cell(row=r, column=c).value), r - 2))
					break
			# assign correct max row value, if any
			self.__max_rows = r - 2
			self.__max_row  = r

			# Iterate through all rows
			act_row = self.__ws.min_row + 1
			act_col = self.__ws.min_column
			diff_prev = timedelta(0)
			for x in range (0, self.__max_rows - 1, 1):
				t1 = self.__ws.cell(row=act_row, column=act_col).value
				t2 = self.__ws.cell(row=act_row + 1, column=act_col).value
				str_t1 = str.format('{}:{}', t1.hour, t1.minute)
				str_t2 = str.format('{}:{}', t2.hour, t2.minute)
				time1 = datetime.strptime(str_t1, self.__cnf['sf_time'])
				time2 = datetime.strptime(str_t2, self.__cnf['sf_time'])
				diff = time2 - time1
				# Compensate for roll over of time to next day
				if (diff.days == -1):
					diff = diff + timedelta(days=1)
				if (diff_prev > timedelta(0) and diff_prev != diff):
					it1 = diff.seconds/60
					it2 = diff_prev.seconds/60
					self.__lgr.error (str.format(self.__cnf['res_strings']['errors']['0014'], it1, self.__ws.cell(act_row, act_col).column_letter, self.__ws.cell(act_row, act_col).row, self.__ws.cell(act_row + 1, act_col).column_letter, self.__ws.cell(act_row + 1, act_col).row, it2))
					raise ValueError (self.__cnf['res_strings']['errors']['0005'].format (self.__cnf['f_frame']))
				diff_prev = diff
				act_row += 1
			# Set start time
			self.__strt_tm = self.__ws.cell(row=self.__ws.min_row + 1, column=act_col).value
			# Set service interval
			self.__svc_intv = diff.total_seconds()/60
			self.__isReady  = True
			self.__isSource = True
			self.__lgr.debug ("{}: {} min, {}: {} hours.".format (self.__cnf['res_strings']['prompts']['0018'], int(self.__svc_intv), self.__cnf['res_strings']['prompts']['0019'], self.__strt_tm.strftime('%H:%M')))
			self.__lgr.debug (str.format(self.__cnf['res_strings']['debug']['0003'], __name__))
			# Check the imported data for validity
			self.xlCheckImportDataValidity()
		except ValueError:
			raise Exception
		except Exception as ex:
			raise (ex)

# ================== Pre-process Excel Source Data Complex -- NOT THREAD SAFE ==================
	def xlPreProcessComplex (self):
		try:
			err_msg = ''
			# load the source forecast data from xl
			self.__lgr.info (str.format(self.__cnf['res_strings']['info']['0038'], self.__wb_fname))
			self.__wb = load_workbook(self.__wb_fname, data_only=True)
			self.__lgr.debug (str.format(self.__cnf['res_strings']['excel']['0003'], self.__wb_fname))

	# Check if there is a framework sheet provided in the Excel workbook and if we should use it
			if self.__env['excel']['useFramework'] and not self.__hasFW:
				# Check if there is a Framework Sheet
				sheets = self.__wb.get_sheet_names()
				for sheet in sheets:
					if sheet == self.__env['excel']['FWS_name']:
						self.__hasFW = True
						self.__lgr.info (str.format(self.__cnf['res_strings']['excel']['0001'], self.__env['excel']['FWS_name'], self.__wb_fname))
				# Yes, there is - let's use it
				if self.__hasFW:
					self.xlReadFrameworkData()
			# Select the active data sheet
			self.__ws = self.__wb[self.__env['excel']['TS_Title']]
		except:
			# Something went wrong
			err = str.format(self.__cnf['res_strings']['errors']['0017'], self.__wb_fname)
			self.__lgr.error (err)
			raise Exception (err)

	# Determine Dates, Times and Service Interval from xl sheet
		self.__max_rows = self.__ws.max_row
		self.__max_cols = self.__ws.max_column
		self.__min_row  = self.__ws.min_row
		self.__min_col  = self.__ws.min_column
		cur_col         = self.__min_col
		col_offset      = self.__env['excel_map']['time'][0]

		try:
			# This Excel spreadsheet is organised as follows:
			# 1st column: Date - repeats n times according to interval and operation hours, i.e. 24 hours, interval 1 hour = 24 entries per day
			# 2nd column: Time - operations hours x interval, i.e. 24 hours operation an 1 hour interval = 24 entries
			# 3rd and subsequent columns - data
			# 1st row - headings 
			# Check that all columns contain the valid format and are not blank
			dateobj_2 = datetime(year=1900, month=1, day=1)
			for r in range (2, self.__max_rows, 1):
				# Seems to be more logical for log-file to check for empty cells first, so do not change the order of the two if statements
				# Check the date first, and while that's happening, lets count the number of days
				if (self.__ws.cell(r, cur_col).value is None):
					self.__ws.cell(r, cur_col).fill = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')
					# Cannot have blank date cells - raise an error
					err_msg = str.format(self.__cnf['res_strings']['excel']['0011'], self.__ws.cell(r, cur_col).column_letter, self.__ws.cell(r, cur_col).row)
					self.__lgr.error (err_msg)
					raise TypeError (err_msg)
				elif (not self.__ws.cell(r, cur_col).is_date):
					self.__ws.cell(r, cur_col).fill = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')
					err_msg = str.format(self.__cnf['res_strings']['excel']['0012'], self.__ws.cell(r, cur_col).column_letter, self.__ws.cell(r, cur_col).row)
					self.__lgr.warning (err_msg)
					raise TypeError (err_msg)
				else:
					dateobj_1 = self.__ws.cell(r, cur_col).value
					if dateobj_1 is not None:
						if dateobj_2 is not None:
							if dateobj_1 != dateobj_2:
								self.__max_days += 1
								self.__date_ary.append(dateobj_1)
						if r > 1:
							dateobj_2 = self.__ws.cell(r, cur_col).value
				# Check the times column
				if (self.__ws.cell(r, col_offset).value is None):
					self.__ws.cell(r, col_offset).fill = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')
					# Cannot have blank date cells - raise an error
					err_msg = str.format(self.__cnf['res_strings']['excel']['0011'], self.__ws.cell(r, col_offset).column_letter, self.__ws.cell(r, col_offset).row)
					self.__lgr.error (err_msg)
					raise TypeError (err_msg)
				elif (not self.__ws.cell(r, col_offset).is_date):
					self.__ws.cell(r, col_offset).fill = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')
					err_msg = str.format(self.__cnf['res_strings']['excel']['0012'], self.__ws.cell(r, cur_col).column_letter, self.__ws.cell(r, col_offset).row)
					self.__lgr.warning (err_msg)
					raise TypeError (err_msg)
				else:
					try:
						self.__time_ary.index(self.__ws.cell(r, col_offset).value)
					except:
						self.__time_ary.append(self.__ws.cell(r, col_offset).value)
						self.__cnt_iv += 1

			# Get the start and end date of the Excel source data
			self.__end_dt	= dateobj_1
			self.__strt_dt	= self.__date_ary[0]
			self.__strt_tm	= self.__time_ary[0]
			# assign correct max row value, if any
			self.__max_rows = r
			act_col = self.__env['excel_map']['time'][0]
			# Iterate through all rows
			diff_prev = timedelta(0)
			for r in range (2, self.__max_rows + 1, 1):
				t1 = self.__ws.cell(row=r, column=act_col).value
				t2 = self.__ws.cell(row=r + 1, column=act_col).value
				str_t1 = str.format('{}:{}', t1.hour, t1.minute)
				str_t2 = str.format('{}:{}', t2.hour, t2.minute)
				time1 = datetime.strptime(str_t1, self.__cnf['sf_time'])
				time2 = datetime.strptime(str_t2, self.__cnf['sf_time'])
				diff = time2 - time1
				# Compensate for roll over of time to next day
				if (diff.days == -1):
					diff = diff + timedelta(days=1)
				if (diff_prev > timedelta(0) and diff_prev != diff):
					it1 = diff.seconds/60
					it2 = diff_prev.seconds/60
					self.__lgr.error (str.format(self.__cnf['res_strings']['errors']['0014'], it1, self.__ws.cell(r, act_col).column_letter, self.__ws.cell(r, act_col).row, self.__ws.cell(r + 1, act_col).column_letter, self.__ws.cell(r + 1, act_col).row, it2))
					raise ValueError (self.__cnf['res_strings']['errors']['0005'].format (self.__cnf['f_frame']))
				diff_prev = diff
			
			# Set service interval
			self.__svc_intv = diff.total_seconds()/60
			self.__isReady  = True
			self.__isSource = True
			self.__lgr.debug (self.__cnf['res_strings']['excel']['0015'].format (self.__cnf['res_strings']['prompts']['0018'], int(self.__svc_intv), self.__cnf['res_strings']['prompts']['0019'], self.__strt_tm.strftime('%H:%M'), self.__cnf['res_strings']['prompts']['0023'], int (self.__max_days), self.__cnf['res_strings']['prompts']['0022'], self.__cnt_iv))
			self.__lgr.debug (str.format(self.__cnf['res_strings']['debug']['0003'], __name__))
			# Check the imported data for validity
			self.xlCheckImportDataValidity()
		except KeyError:
			raise KeyError (self.__cnf['res_strings']['errors']['0020'])
		except Exception as ex:
			raise (ex)
		finally:
			del err_msg

# ================== Construct the forecast dictionary  -- NOT THREAD SAFE -- ==================
	def xlCreateDictionary (self):
		"""
		Function creates the dictionary used for forecasting the agent numbers.
		
		Parameters
		----------
		dic_fc : dictionary dict {}
			Generates the excel reports based on forecast data provided in fc.
			The report parameters are defined in file workforce.conf
		"""

		try:
			if (self.__isReady):
				# Create the array for the various times - the size depends on the forecast interval 15, 30, 45, 60 minutes and the operations hours
				# The times array is common to all other entries in the folowing arrays, so don't need to duplicate that for each Dayx
				self.__dic_fc['times'] = []
				# Use start time from before
				timeobj = self.__strt_tm
				# ================== Construct the forecast dictionary from Excel data ==================
				self.__lgr.debug  (self.__cnf['res_strings']['info']['0007'])
				# First creating the Interval times array

				for c in range (0, self.__cnt_iv, 1):
					t2 = self.__ws.cell(row=self.__ws.min_row + c + 1, column=self.__ws.min_column).value
					str_t2 = str.format('{}:{}', t2.hour, t2.minute)
					timeobj = datetime.strptime(str_t2, self.__cnf['sf_time'])
					self.__dic_fc['times'].append(timeobj.strftime (self.__cnf['sf_time']))
				self.__lgr.debug (str.format(self.__cnf['res_strings']['info']['0002'], len(self.__dic_fc['times'])))
				# Create the rest of the arrays, all but Calls will be empty for now
				if (self.__env['export']['report-detail']):
					self.__lgr.debug (self.__cnf['res_strings']['debug']['0004'])
				else:
					self.__lgr.debug (self.__cnf['res_strings']['debug']['0005'])
				for i in range (0, self.__max_days, 1): # restrict to maximum forecast days as calculated before
					# Read date from excel spread sheet and put into array
					row_num = self.__ws.min_row
					col_num = self.__ws.min_column + i + 1
					dateobj = self.__ws.cell(row=row_num, column=col_num).value
					self.__lgr.debug (str.format (self.__cnf['res_strings']['info']['0001'], dateobj.strftime (self.__cnf['sf_date'])))
					# Format the dictionary key - Set0, Set1, Set2, ...
					str_tmp = 'Day' + str(i) 
					self.__dic_fc[str_tmp] = {}
					self.__dic_fc[str_tmp]['count'] = self.__max_days
					self.__dic_fc[str_tmp]['date'] = dateobj.strftime (self.__cnf['sf_date'])
					# Insert Call Numbers for this date
					self.__dic_fc[str_tmp]['calls'] = []
					for c in range(0, self.__cnt_iv):
						r = c + 2
						val = self.__ws.cell(row=r, column=col_num).value
						self.__dic_fc[str_tmp]['calls'].append(int(val))
					self.__dic_fc[str_tmp]['agents'] = []
					if (self.__env['export']['report-detail']):
						self.__dic_fc[str_tmp]['util']      = []
						self.__dic_fc[str_tmp]['sla']       = []
						self.__dic_fc[str_tmp]['asa']       = []
						self.__dic_fc[str_tmp]['abandon']   = []
						self.__dic_fc[str_tmp]['q-percent'] = []
						self.__dic_fc[str_tmp]['q-time']    = []
						self.__dic_fc[str_tmp]['q-count']   = []
				# Signal the the dictionary has been built and copied
				self.__hasDict = True
				self.__lgr.debug (self.__cnf['res_strings']['info']['0008'])
				return self.__dic_fc
			else:
				raise RuntimeError (str.format(self.__cnf['res_strings']['errors']['0016']))
		except Exception as ex:
			raise (ex)

# ================== Sanity check for data read - compare parts of framework file with actual data -- NOT THREAD SAFE -- ==================
	def xlCheckImportDataValidity (self):
		try:
			if (self.__isReady):
				# Check if Service Interval is 15, 30, 45 or 60 minutes
				if (self.__svc_intv) not in self.__cnf['fw']['Intervals']:
					raise ValueError (self.__cnf['res_strings']['errors']['0001'].format(self.__svc_intv, self.__wb_fname))
				# Check if Service Interval from framework file is the same as in excel file
				if (self.__svc_intv != self.__cnf['fw']['ServiceInterval'] or self.__cnf['fw']['ServiceInterval'] <= 0):
					raise ValueError (self.__cnf['res_strings']['errors']['0002'].format (self.__cnf['fw']['ServiceInterval'], self.__cnf['f_frame'], int(self.__svc_intv), self.__wb_fname))
				if self.__cnf['fw']['OperationHours'] <= 0:
					raise ValueError (str.format(self.__cnf['res_strings']['errors']['0006'], self.__cnf['fw']['OperationHours'], self.__cnf['f_frame']))

				# Calculate maximum Number of Intervals per calculation loop
				self.__cnt_iv = int (self.__cnf['fw']['OperationHours'] * (60 / self.__svc_intv))
				# Check if framework interval is different to what is in the excel forcast sheet
				if self.__env['excel']['Simple']:
					if (self.__cnt_iv != self.__max_rows):
						self.__lgr.warning (self.__cnf['res_strings']['errors']['0004'])
						self.__cnt_iv = int(self.__max_rows)
					# Check forecast framework forcast days against spreadsheet
					if ((self.__max_cols) != self.__cnf['fw']['ForecastDays']):
						# There are less forecast columns than forecast days specified in the framework file, using the columns available
						if ((self.__max_cols) <= self.__cnf['fw']['ForecastDays']):
							self.__lgr.warning (str.format(self.__cnf['res_strings']['warnings']['0001'], self.__max_cols, self.__wb_fname, self.__cnf['fw']['ForecastDays'], self.__cnf['f_frame']))
							self.__max_days = self.__max_cols
						# There are more columns in the forecast file than specified in the framework file - using framework file value
						if ((self.__max_cols) >= self.__cnf['fw']['ForecastDays']):
							self.__lgr.warning (str.format(self.__cnf['res_strings']['warnings']['0001'], self.__max_cols, self.__wb_fname, self.__cnf['fw']['ForecastDays'], self.__cnf['f_frame']))
							self.__max_days , self.__cnf['fw']['ForecastDays']
					else:
						self.__max_days = self.__max_cols
				else:
					if len (self.__time_ary) != int (60 / self.__cnf['fw']['ServiceInterval'] * self.__cnf['fw']['OperationHours']):
						self.__lgr.warning (str.format(self.__cnf['res_strings']['warnings']['0001'], self.__max_days, self.__wb_fname, self.__cnf['fw']['ForecastDays'], self.__cnf['f_frame']))
				self.__lgr.info (str.format(self.__cnf['res_strings']['info']['0005'], self.__max_days))
				self.__lgr.info (str.format(self.__cnf['res_strings']['info']['0006'], self.__cnt_iv))
				self.__lgr.info (str.format(self.__cnf['res_strings']['info']['0004'], self.__cnt_iv * self.__max_days))
				self.__lgr.info (self.__cnf['res_strings']['excel']['0017'])
				return
			else:
				raise RuntimeError (str.format(self.__cnf['res_strings']['errors']['0016']))
		except Exception as ex:
			raise (ex)

# ================== Create the Excel report -- THREAD SAFE -- ==================
	def xlCreateReport (self, fc:dict): 
		"""
		Function creates agent number forecast based on data provided.
		The forecast is exported to an Excel (.xlsx). Forecast outcome is based on
		parameters provided during class initialisation.

		Parameters
		----------
		fc : custom dictionary
			Generates the excel reports based on forecast data provided in fc.
			The report parameters are defined in file workforce.conf
		"""
		try:
			if (self.__isReady):
			# Load the forecast output excel file
				self.__wb = Workbook()
				self.__wb.save (self.__cnf['f_result'])
				self.__wb = load_workbook (self.__cnf['f_result'])
				self.__lgr.info (str.format(self.__cnf['res_strings']['info']['0015'], self.__cnf['f_result']))
				self.__isSource = False                
				# Create the summary worksheet
				# wb.create_sheet (title=env['excel']['summarysheet'])
				# Select the active sheet
				self.__ws = self.__wb.active
				self.__ws.title = self.__env['excel']['SS_name']
				self.__lgr.debug (str.format(self.__cnf['res_strings']['info']['0022'], self.__ws.title))
				self.__wb.save (self.__cnf['f_result'])
				self.__lgr.info (self.__cnf['res_strings']['info']['0028'])

			# Start filling in the details - starting with the times
				cell = self.__ws.cell(1,1) # 'tis cell 'A1'
				cell.value = 'Times'
				times = fc['times']
				row = 2
				col = 1
				for time in times:
					self.__ws.cell(row, col).value = time
					row += 1
				self.__lgr.debug (self.__cnf['res_strings']['info']['0023'])

			# Next fill the dates in the first row and at the same time create the various detail sheets if required
				s_day = ''
				i = 0
				row = 1
				col = 2
				for i in range(len(fc) - 1):
					s_day = 'Day' + str(i)
					self.__ws.cell(row, col).value = fc[s_day]['date']
					# Only create detail excel sheets and report if configured
					if (self.__env['export']['report-detail']):
						dateobj = datetime.strptime(fc[s_day]['date'], self.__cnf['sf_date'])
						sheet_name = dateobj.strftime(self.__env['formats']['xl-detail'])
						self.__wb.create_sheet (sheet_name)
						self.__lgr.debug (str.format(self.__cnf['res_strings']['debug']['0002'], sheet_name))
					self.__ws = self.__wb[self.__env['excel']['SS_name']]
					i += 1
					col += 1
				self.__lgr.debug (self.__cnf['res_strings']['info']['0024'])

			# Fill in the agents numbers into the summary sheetrequired 
				col = 2
				for i in range(len(fc) - 1):
					row = 2
					s_day = 'Day' + str(i)
					agents = fc[s_day]['agents']
					for agent in agents:
						self.__ws.cell(row, col).value = agent
						row += 1
					i += 1
					col += 1
				self.__lgr.info (self.__cnf['res_strings']['info']['0025'])

			# Create detailed report - if required
				if (self.__env['export']['report-detail']):
					self.__lgr.info (self.__cnf['res_strings']['info']['0026'])
					self.__max_sht = len(self.__wb.sheetnames) - 1
					# Iterate through all sheets in workbook 
					day = 0
					for self.__ws in self.__wb:
						row = 1
						col = 1
						# Skip the summary sheet
						if (self.__ws.title == self.__env['excel']['SS_name']):
							continue
						# Fill the heading for each sheet
						for head in self.__env['export']['headings']:
							self.__ws.cell(1, col, value=head)
							col += 1
						# Fill spreadsheet with times and other details
						times = fc['times']
						x = 0
						r = 2
						s_day = 'Day' + str(day)
						fc_day = self.__dic_fc[s_day]
						for time in times:
							self.__ws.cell(r, column=1).value = time
							self.__ws.cell(r, column=2).value = fc_day['calls'][x]
							self.__ws.cell(r, column=3).value = fc_day['agents'][x]
							self.__ws.cell(r, column=4).value = fc_day['sla'][x]
							self.__ws.cell(r, column=4).number_format = self.__env['formats']['percent']
							self.__ws.cell(r, column=5).value = fc_day['asa'][x]
							self.__ws.cell(r, column=6).value = fc_day['abandon'][x]
							self.__ws.cell(r, column=6).number_format = self.__env['formats']['percent']
							self.__ws.cell(r, column=7).value = fc_day['q-percent'][x]
							self.__ws.cell(r, column=7).number_format = self.__env['formats']['percent']
							self.__ws.cell(r, column=8).value = fc_day['q-time'][x]
							self.__ws.cell(r, column=9).value = fc_day['q-count'][x]
							x += 1
							r += 1
						day += 1
				else:
					self.__lgr.info (self.__cnf['res_strings']['info']['0027'])

			# Exit gracefully
				self.__wb.save (self.__cnf['f_result'])
				self.__lgr.info (str.format(self.__cnf['res_strings']['info']['0011'], self.__cnf['f_result']))
				return
			else:
				raise RuntimeError (str.format(self.__cnf['res_strings']['errors']['0016']))
		except Exception as ex:
			raise (ex)

# ================== Function to read Framework Data from Excel File -- NOT THREAD SAFE -- ==================
	def xlReadFrameworkData (self):
		"""
		Function reads the framework data used for the forecasting process from Excel Workbook
		and checks it for errors.
		
		Parameters
		----------
		None
		"""
		# Check if the framework info has been loaded - source can be framework json or excel
		if not self.__hasFW:
			return
		fw = {}
		try:
			fw = self.__cnf['fw']
			ws = self.__wb[self.__env['excel']['FWS_name']]
			col = ws.min_column
			# Iterate through the rows in the spreadsheet
			self.__lgr.info (str.format(self.__cnf['res_strings']['excel']['0002'], self.__wb_fname, self.__env['excel']['FWS_name']))
			for row in range (ws.min_row, ws.max_row + 1, 1):
				# skip any empty cells
				if ws.cell(row, col).value is None:
					continue
				# Map the names and values in excel to the framework dictionary
				for k in fw.keys():
					# Check if this is the array that we are expecting ('Intervals')
					if ws.cell(row, col).value == k:
						if type (fw[k]).__name__ in ('list', 'tuple'):
							if ws.cell (row, col + 1).value is not None: fw[k][0] = ws.cell (row, col + 1).value
							if ws.cell (row, col + 2).value is not None: fw[k][0] = ws.cell (row, col + 2).value
							if ws.cell (row, col + 3).value is not None: fw[k][0] = ws.cell (row, col + 3).value
							if ws.cell (row, col + 4).value is not None: fw[k][0] = ws.cell (row, col + 4).value
						else:
							fw[k] = ws.cell (row, col + 1).value
						break
			wfm_helpers.helperCheckFrameworkData (self.__cnf, self.__env, self.__lgr)
			# Signal that the framework parameters have been initialised from Excel.
			self.__hasFW = True   
		except Exception as e:
			self.__lgr.info (self.__cnf['res_strings']['excel']['0004'])
			raise (e)

# ================== Helper Function to create a record using call volumes only for export to DB -- NOT THREAD SAFE -- ==================
	def xlImportTxnsOnly (self, record:dict, col=0):
		try:
			# Make sure the Excel Spreadsheet has been processed and checked before continuing
			# But only do this once, as the class will keep the important values in local variables
			if not self.__isReady or not self.__isSource:
				# Simple means we only are importing transaction data, times and dates - nothing else
				# it means that data is organised:
				#       Date,   Date,   Date,   ...
				# Time  txns,   txns,   txns,   ...
				# Time  txns,   txns,   txns,   ...
				# Time  txns,   txns,   txns,   ...
				self.xlPreProcessSimple ()
			sum_txn         = 0
			date_cell       = self.__ws.cell(row=1, column=col).value
			record['_id']   = int (str (datetime.strftime (date_cell, self.__env['formats']['dbq-date'])))
			record['day_ary'][0]['txn_sum']['year']    = int (date_cell.year)
			record['day_ary'][0]['txn_sum']['month']   = int (date_cell.month)
			record['day_ary'][0]['txn_sum']['day']     = int (date_cell.day)
			record['day_ary'][0]['txn_sum']['s_mnth']  = str (datetime.strftime (date_cell, '%B'))
			record['day_ary'][0]['txn_sum']['s_date']  = str (datetime.strftime (date_cell, self.__env['formats']['dbq-date']))
			record['day_ary'][0]['txn_sum']['dow']     = str (datetime.strftime (date_cell, '%A'))

			# TODO - create the 'xl_db_map dynamically from spreadsheet
			if not self.__env['excel']['Simple']:
				inc = -2 # account for the three keys that always need to be there, otherwise we have nothing to import
				# Calculate the increment based on how many columns of KPIs there are
				tmp_dict = {}
				tmp_dict = self.__env['excel_map']
				for k, v in tmp_dict.items():
					if v[2]:
						inc += 1
			for r in range (2, self.__max_row, 1):
				t1 = self.__ws.cell(row=r, column=1).value
				record['day_ary'][0]['times'].append (str.format ('{:02d}:{:02d}', t1.hour, t1.minute))
				val = int (round (self.__ws.cell(row=r, column=col).value))
				record['day_ary'][0]['tx_all'].append (val)
				sum_txn += val

			# Sum of all transactions received
			record['day_ary'][0]['txn_sum']['sum_txn'] = sum_txn
		except Exception as ex:
			raise (ex)

# ================== Helper Function to create a record for export all KPIs to DB -- NOT THREAD SAFE -- ==================
	def xlImportTxnsKPIs (self, record:dict, date=''):
		try:
			# Make sure the Excel Spreadsheet has been processed and checked before continuing
			# But only do this once, as the class will keep the important values in local variables
			if not self.__isReady or not self.__isSource:
				# Complex means we are importing all data specified in framework.conf.excel_map
				# it also means, that the data is organised:
				# date, time, metric 1, metric 2, ... 
				# date, time, metric 1, metric 2, ... 
				# date, time, metric 1, metric 2, ... 
				self.xlPreProcessComplex ()

			val_txn = 0
			val_ans = 0
			val_iat = 0
			val_aiw = 0
			sum_txn	= 0
			sum_ans	= 0
			sum_abn	= 0
			sum_at	= 0
			sum_aiw	= 0
			sum_ht	= 0
			sum_lit	= 0
			sum_rt	= 0
			sum_ag	= 0
			col		= self.__env['excel_map']['date'][0]
			# Need to add start date check for start date plus max days
			if date is not None:
				if type (date) is not datetime:
					raise TypeError (self.__cnf['res_strings']['excel']['0019'].format (self.__class__, 'xlImportTxnsKPIs'))
				if date < self.__strt_dt:
					raise ValueError (self.__cnf['res_strings']['excel']['0016'].format (self.__class__, 'xlImportTxnsKPIs', self.__strt_dt, self.__wb_fname))
				elif date > self.__end_dt:
					raise ValueError (self.__cnf['res_strings']['excel']['0018'].format (self.__class__, 'xlImportTxnsKPIs', self.__end_dt, self.__wb_fname))
				else:    
					date_cell = date
			else:
				date_cell = self.__strt_dt
			record['_id'] = int (str (datetime.strftime (date_cell, self.__env['formats']['dbq-date'])))
			record['day_ary'][0]['txn_sum']['year']    = int (date_cell.year)
			record['day_ary'][0]['txn_sum']['month']   = int (date_cell.month)
			record['day_ary'][0]['txn_sum']['day']     = int (date_cell.day)
			record['day_ary'][0]['txn_sum']['s_mnth']  = str (datetime.strftime (date_cell, '%B'))
			record['day_ary'][0]['txn_sum']['s_date']  = str (datetime.strftime (date_cell, self.__env['formats']['dbq-date']))
			record['day_ary'][0]['txn_sum']['dow']     = str (datetime.strftime (date_cell, '%A'))

			# Need to iterate to the start of the sequence given by date_cell
			start_row	= 2
			col			= self.__env['excel_map']['date'][0]

			for start_row in range (2, self.__max_rows, 1):
				if date != self.__ws.cell(row=start_row, column=col).value:
					continue
				else:
					self.__row_offs += 1
					break
			# Retrieve data from spreadsheet
			for r in range (start_row, self.__cnt_iv * self.__row_offs + 1, 1):
				# Time column
				col = self.__env['excel_map']['time'][0]
				t1 = self.__ws.cell(row=r, column=col).value
				record['day_ary'][0]['times'].append (str.format ('{:02d}:{:02d}', t1.hour, t1.minute))
				
				# Get all transactions offered: tx_all
				col = self.__env['excel_map']['tx_all'][0]
				val_txn = int (self.__ws.cell(row=r, column=col).value)
				if val_txn < 0: 
					raise ValueError (self.__cnf['res_strings']['excel']['0020'].format (self.__ws.cell(row=r, column=col).column_letter, self.__ws.cell(row=r, column=col).row))
				record['day_ary'][0]['tx_all'].append (val_txn)
				# Add to sum of all transaction for that day
				sum_txn += val_txn

			# Process additional columns according to 'xl_db_map' from config file
				# Transactions handled
				if self.__env['excel_map']['tx_ans'][1]:
					col = self.__env['excel_map']['tx_ans'][0]
					val_ans = int (self.__ws.cell(row=r, column=col).value)
					if val_ans < 0: 
						raise ValueError (self.__cnf['res_strings']['excel']['0020'].format (self.__ws.cell(row=r, column=col).column_letter, self.__ws.cell(row=r, column=col).row))
					record['day_ary'][0]['tx_ans'].append (val_ans)
				else:
					val_ans = val_txn	
				sum_ans += val_ans

				# Transactions Abandoned
				if self.__env['excel_map']['tx_abn'][1]:
					col = self.__env['excel_map']['tx_abn'][0]
					val_abn = int (self.__ws.cell(row=r, column=col).value)
					if val_abn < 0: 
						raise ValueError (self.__cnf['res_strings']['excel']['0020'].format (self.__ws.cell(row=r, column=col).column_letter, self.__ws.cell(row=r, column=col).row))
					record['day_ary'][0]['tx_abn'].append (val_abn)
				else:
					val_abn = val_txn - val_ans	
				sum_abn += val_abn

				# Abandonment Rate
				if self.__env['excel_map']['r_abn'][1]:
					col = self.__env['excel_map']['r_abn'][0]
					if self.__ws.cell(row=r, column=col).value < 0: 
						raise ValueError (self.__cnf['res_strings']['excel']['0020'].format (self.__ws.cell(row=r, column=col).column_letter, self.__ws.cell(row=r, column=col).row))
					record['day_ary'][0]['r_abn'].append (round (self.__ws.cell(row=r, column=col).value, 2))
				else:
					if val_ans > 0:
						record['day_ary'][0]['r_abn'].append (round ((val_txn - val_ans) / val_txn, 2))
				sum_abn += (val_txn - val_ans)
				# Service Level
				if self.__env['excel_map']['sl'][1]:
					col = self.__env['excel_map']['sl'][0]
					if self.__ws.cell(row=r, column=col).value < 0: 
						raise ValueError (self.__cnf['res_strings']['excel']['0020'].format (self.__ws.cell(row=r, column=col).column_letter, self.__ws.cell(row=r, column=col).row))
					record['day_ary'][0]['sl'].append (self.__ws.cell(row=r, column=col).value)
				# Average Speed to Answer
				if self.__env['excel_map']['a_sa'][1]:
					col = self.__env['excel_map']['a_sa'][0]
					if self.__ws.cell(row=r, column=col).value < 0: 
						raise ValueError (self.__cnf['res_strings']['excel']['0020'].format (self.__ws.cell(row=r, column=col).column_letter, self.__ws.cell(row=r, column=col).row))
					record['day_ary'][0]['a_sa'].append (self.__ws.cell(row=r, column=col).value)
				# Average Wait Time
				if self.__env['excel_map']['a_wait'][1]:
					col = self.__env['excel_map']['a_wait'][0]
					if self.__ws.cell(row=r, column=col).value < 0: 
						raise ValueError (self.__cnf['res_strings']['excel']['0020'].format (self.__ws.cell(row=r, column=col).column_letter, self.__ws.cell(row=r, column=col).row))
					record['day_ary'][0]['a_wait'].append (self.__ws.cell(row=r, column=col).value)
				# Average Queue Rate
				if self.__env['excel_map']['r_queue'][1]:
					col = self.__env['excel_map']['r_queue'][0]
					if self.__ws.cell(row=r, column=col).value < 0: 
						raise ValueError (self.__cnf['res_strings']['excel']['0020'].format (self.__ws.cell(row=r, column=col).column_letter, self.__ws.cell(row=r, column=col).row))
					record['day_ary'][0]['q-rate'].append (self.__ws.cell(row=r, column=col).value)
				# Incident Active Times
				if self.__env['excel_map']['iat'][1]:
					col = self.__env['excel_map']['iat'][0]
					val = self.__ws.cell(row=r, column=col).value
					if val < 0: 
						raise ValueError (self.__cnf['res_strings']['excel']['0020'].format (self.__ws.cell(row=r, column=col).column_letter, self.__ws.cell(row=r, column=col).row))
					record['day_ary'][0]['iat'].append (val)
					sum_at += val
				# After Incident Wrap Times
				if self.__env['excel_map']['aiw'][1]:
					col = self.__env['excel_map']['aiw'][0]
					val = self.__ws.cell(row=r, column=col).value
					if val < 0: 
						raise ValueError (self.__cnf['res_strings']['excel']['0020'].format (self.__ws.cell(row=r, column=col).column_letter, self.__ws.cell(row=r, column=col).row))
					record['day_ary'][0]['aiw'].append (val)
					sum_aiw += val
				# Total Incident Times
				if self.__env['excel_map']['iht'][1]:
					col = self.__env['excel_map']['iht'][0]
					val = self.__ws.cell(row=r, column=col).value
					if val < 0: 
						raise ValueError (self.__cnf['res_strings']['excel']['0020'].format (self.__ws.cell(row=r, column=col).column_letter, self.__ws.cell(row=r, column=col).row))
					record['day_ary'][0]['iht'].append (val)
					sum_ht += val
				# Agent Numbers
				if self.__env['excel_map']['agents'][2]:
					col = self.__env['excel_map']['agents'][0]
					val = self.__ws.cell(row=r, column=col).value
					if val < 0: 
						raise ValueError (self.__cnf['res_strings']['excel']['0020'].format (self.__ws.cell(row=r, column=col).column_letter, self.__ws.cell(row=r, column=col).row))
					record['day_ary'][0]['agents'].append (val)
					sum_ag += val
				# Agent Login Times
				if self.__env['excel_map']['ag_lit'][2]:
					col = self.__env['excel_map']['ag_lit'][0]
					val = self.__ws.cell(row=r, column=col).value
					if val < 0: 
						raise ValueError (self.__cnf['res_strings']['excel']['0020'].format (self.__ws.cell(row=r, column=col).column_letter, self.__ws.cell(row=r, column=col).row))
					record['day_ary'][0]['ag_lit'].append (val)
					sum_lit += val
				# Agent Rostered Times
				if self.__env['excel_map']['ag_rt'][2]:
					col = self.__env['excel_map']['ag_rt'][0]
					val = self.__ws.cell(row=r, column=col).value
					if val < 0: 
						raise ValueError (self.__cnf['res_strings']['excel']['0020'].format (self.__ws.cell(row=r, column=col).column_letter, self.__ws.cell(row=r, column=col).row))
					record['day_ary'][0]['ag_rt'].append (val)
					sum_rt += val
				# Utilisation Rate (Sum AHT / Logged-in Time)
				if self.__env['excel_map']['r_util'][2]:
					col = self.__env['excel_map']['r_util'][0]
					if val < 0: 
						raise ValueError (self.__cnf['res_strings']['excel']['0020'].format (self.__ws.cell(row=r, column=col).column_letter, self.__ws.cell(row=r, column=col).row))
					record['day_ary'][0]['r_util'].append (self.__ws.cell(row=r, column=col).value)
				# Agent Availability (Logged-in Time / Rostered Time)
				if self.__env['excel_map']['r_avail'][2]:
					col = self.__env['excel_map']['r_avail'][0]
					if val < 0: 
						raise ValueError (self.__cnf['res_strings']['excel']['0020'].format (self.__ws.cell(row=r, column=col).column_letter, self.__ws.cell(row=r, column=col).row))
					record['day_ary'][0]['r_avail'].append (self.__ws.cell(row=r, column=col).value)

		# Create summary and average values
			# Sum of all transactions received
			record['day_ary'][0]['txn_sum']['sum_txn'] = sum_txn
			# Sum of all transactions answered
			record['day_ary'][0]['txn_sum']['sum_ans'] = sum_ans
			# Calculate sum of all transactions abandoned
			record['day_ary'][0]['txn_sum']['sum_abn'] = sum_txn - sum_ans
			# Calculate abandonment rate for whole day
			if sum_txn > 0:
				record['day_ary'][0]['txn_sum']['r_abn'] = round ((sum_txn - sum_ans) / sum_txn, 2)
			# Calculate Total Handle, After Incident Wrap and Active Time
			if sum_ans > 0:
				if sum_ht <= 0:
					record['day_ary'][0]['txn_sum']['sum_ht'] = sum_at + sum_aiw
				else:
					record['day_ary'][0]['txn_sum']['sum_ht'] = sum_ht
				record['day_ary'][0]['txn_sum']['sum_lit']	= sum_lit
				record['day_ary'][0]['txn_sum']['sum_rt']	= sum_rt
				record['day_ary'][0]['txn_sum']['sum_at']	= sum_at
				record['day_ary'][0]['txn_sum']['sum_aiw']	= sum_aiw
				record['day_ary'][0]['txn_sum']['avg_ht']	= int (round ((sum_at + sum_aiw) / sum_ans, 0))
				record['day_ary'][0]['txn_sum']['avg_at']	= int (round (sum_at / sum_ans, 0))
				record['day_ary'][0]['txn_sum']['avg_aiw']	= int (round (sum_aiw / sum_ans, 0))
				# Average Agent Count
				if sum_ag > 0 and self.__cnt_iv > 0:
					record['day_ary'][0]['txn_sum']['avg_ag'] = round (sum_ag / self.__cnt_iv, 2)
				# Agent Availability
				if sum_lit > 0 and sum_rt > 0:
					record['day_ary'][0]['txn_sum']['r_avail'] = round (sum_lit / sum_rt, 2)
				# Agent Utilisation
				if sum_lit > 0 and sum_ht > 0:
					record['day_ary'][0]['txn_sum']['r_util'] = round (sum_lit / sum_ht, 2)
		except KeyError:
			self.__lgr.error (self.__cnf['res_strings']['errors']['0020'])
			raise self.__cnf['res_strings']['errors']['0020']
		except Exception as ex:
			raise (ex)

# ================== Create Excel Template for import of historic data -- THREAD SAFE -- ==================
	def xlCreateImportTemplate (self):
		try:
			if os.path.exists (self.__cnf['f_template']) == False:
				wb = Workbook()
				wb.save (self.__cnf['f_template'])                
				self.__lgr.info (str.format(self.__cnf['res_strings']['info']['0019'], self.__cnf['f_template']))
			self.__lgr.info (str.format (self.__cnf['res_strings']['excel']['0005'], self.__cnf['f_template']))
			wb = load_workbook (self.__cnf['f_template'])
			self.__lgr.debug (str.format (self.__cnf['res_strings']['excel']['0003'], self.__cnf['f_template']))

			# Delete any existing worksheets i the workbook
			sheets = wb.sheetnames
			self.__lgr.info (str.format (self.__cnf['res_strings']['excel']['0006'], self.__cnf['f_template']))
			for name in sheets:
				ws = wb.get_sheet_by_name(name)
				wb.remove(ws)
			# Add one worksheet that represents the format needed for Complex Data import
			self.__lgr.info (str.format (self.__cnf['res_strings']['excel']['0007']))
			wb.create_sheet (self.__env['excel']['TS_Title'])
			ws = wb[self.__env['excel']['TS_Title']]
			# Create Heading Row
			self.__lgr.info (str.format (self.__cnf['res_strings']['excel']['0008']))
			for key in self.__env['excel_map']:
				if key == ("__map-key"):
					continue
				ary = self.__env['excel_map'][key]
				ws.cell(1, ary[0]).value = ary[2]
				print ("Column: {}".format (chr(ary[0] + 64)))

			self.__lgr.info (str.format (self.__cnf['res_strings']['excel']['0009']))
			wb.save (self.__cnf['f_template'])
			wb.close()
			del wb
			del ws
		except Exception as e:
			self.__lgr.error (e)
			raise (e)