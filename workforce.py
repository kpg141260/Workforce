# workforce.py
# Copyright 2020 by Peter Gossler. All rights reserved.
# Version 0.1.0

import logging
import os
import json
from datetime import datetime, timedelta
from openpyxl import load_workbook
from lib.erlang.erlang_base import Erlang_Base
from lib.erlang.erlang_c import Erlang
import lib.gui.workforce_gui

import wfm_helpers

try:
# >>> Tinkering ===========
    #quickWindow()

# ================== Preparation to get everything going ==================
    # Load configuration file
    f_conf = os.path.join ('conf', 'workforce.conf')
    with open (f_conf) as json_file:
        env = json.load(json_file)
    del f_conf

    # Enable Debug level logging and set logging defaults
    logger_wfm = wfm_helpers.config_logger (env, 'wfm')
    # Create resource dictionary
    dic_cnf    = {'f_summary':'', 'f_result':'','fs_nopath':'', 'fr_nopath':'','f_frame':'', 'f_xl_data':'','sf_time':'','sf_date':'', 'res_strings':{}, 'fw':{}}
    wfm_helpers.createFileNames(dic_cnf, env, logger_wfm)
    wfm_helpers.logForecastFramework (dic_cnf, logger_wfm)

# TODO: #2 add some code to pop up a confirmation window here at a later state
    #if (len(wb.sheetnames) > 1):
    #    pass

# ================== Load and process source data - change here if database is source ==================
    # TODO: #3 Add code here to load data from database
    # load the source forecast data from xl
    wb = load_workbook (dic_cnf['f_xl_data'], data_only=True)
    logger_wfm.info (f"Excel workbook {dic_cnf['f_xl_data']} opened.")

    # Select the active sheet
    ws = wb[env['excel']['SD_name']]
    # Determine the row and column count
    max_rows = ws.max_row
    max_cols = ws.max_column
    # Determine the Service Interval from xl sheet
    act_row = ws.min_row + 1
    act_col = ws.min_column
    x = 0
    diff_prev = timedelta(0)
    # Iterate through all rows
    while x < ws.max_row - 2:
        if (ws.cell(row=act_row, column=act_col).is_date):
            t1 = ws.cell(row=act_row, column=act_col).value
            t2 = ws.cell(row=act_row + 1, column=act_col).value
            str_t1 = str.format('{}:{}', t1.hour, t1.minute)
            str_t2 = str.format('{}:{}', t2.hour, t2.minute)
            time1 = datetime.strptime(str_t1, dic_cnf['sf_time'])
            time2 = datetime.strptime(str_t2, dic_cnf['sf_time'])
            diff = time2 - time1
            # Compensate for roll over of time to next day
            if (diff.days == -1):
                diff = diff + timedelta(days=1)
            if (diff_prev > timedelta(0) and diff_prev != diff):
                raise ValueError (dic_cnf['res_strings']['errors']['0005'].format(dic_cnf['f_frame']))
            diff_prev = diff
            act_row += 1
            x += 1
        else:
            raise ValueError (f'Cell type of cell row: {act_row}, column: {act_col} is not formatted as DateTime.')
    # Set start time and service interval
    start_time = ws.cell(row=ws.min_row + 1, column=act_col).value
    ServiceInterval = diff.total_seconds()/60

# ================== Sanity check for data read - compare parts of framework file with actual data  ==================
    # Check if Service Interval is 15, 30, 45 or 60 minutes
    if (ServiceInterval) not in dic_cnf['fw']['Intervals']:
        raise ValueError (dic_cnf['res_strings']['errors']['0001'].format(ServiceInterval, dic_cnf['f_xl_data']))
    # Check if Service Interval from framework file is the same as in excel file
    if (ServiceInterval != dic_cnf['fw']['ServiceInterval'] or dic_cnf['fw']['ServiceInterval'] <= 0):
        raise ValueError ((dic_cnf['res_strings']['errors']['0002'].format(dic_cnf['fw']['ServiceInterval'], dic_cnf['f_frame'], int(ServiceInterval), dic_cnf['f_xl_data'])))
    if (dic_cnf['fw']['OperationHours'] <= 0):
        raise ValueError (str.format(dic_cnf['res_strings']['errors']['0006'], dic_cnf['fw']['OperationHours'], dic_cnf['f_frame']))

    # Calculate maximum Number of Intervals per calculation loop
    cnt_Interval = int(dic_cnf['fw']['OperationHours'] * (60 / dic_cnf['fw']['ServiceInterval']))
    # Check if framework interval is different to what is in the excel forcast sheet
    if (cnt_Interval != max_rows - 1):
        logger_wfm.warning (dic_cnf['res_strings']['errors']['0004'])
        cnt_Interval = int(max_rows - 1)
    # Check forecast framework forcast days against spreadsheet
    if ((max_cols - 1) != dic_cnf['fw']['ForecastDays']):
        # There are less forecast columns than forecast days specified in the framework file, using the columns available
        if ((max_cols - 1) < dic_cnf['fw']['ForecastDays']):
            logger_wfm.warning (str.format(dic_cnf['res_strings']['warnings']['0001'], max_cols - 1, dic_cnf['f_xl_data'], dic_cnf['fw']['ForecastDays'], dic_cnf['f_frame'], max_cols - 1))
            max_days = max_cols - 1
        # There are more columns in the forecast file than specified in the framework file - using framework file value
        if ((max_cols - 1) > dic_cnf['fw']['ForecastDays']):
            logger_wfm.warning (str.format(dic_cnf['res_strings']['warnings']['0001'], max_cols - 1, dic_cnf['f_xl_data'], dic_cnf['fw']['ForecastDays'], dic_cnf['f_frame'], dic_cnf['fw']['ForecastDays']))
            max_days = dic_cnf['fw']['ForecastDays']
    else:
        max_days = max_cols - 1
    logger_wfm.debug (str.format(dic_cnf['res_strings']['info']['0005'], max_days))
    logger_wfm.debug (str.format(dic_cnf['res_strings']['info']['0006'], cnt_Interval))

# ================== Build the Dictionary template for the required calculations ==================
    str_tmp = ''
    dic_fc = {}
    # Create the array for the various times - the size depends on the forecast interval 15, 30, 45, 60 minutes and the operations hours
    # The times array is common to all other entries in the folowing arrays, so don't need to duplicate that for each Dayx
    dic_fc['times'] = []
    # Use start time from before
    timeobj = time1

# ================== Construct the forecast dictionary ==================
    logger_wfm.info (dic_cnf['res_strings']['info']['0007'])
    # First creating the Interval times array
    for c in range (0, cnt_Interval, 1):
        t2 = ws.cell(row=ws.min_row + c + 1, column=ws.min_column).value
        str_t2 = str.format('{}:{}', t2.hour, t2.minute)
        timeobj = datetime.strptime(str_t2, dic_cnf['sf_time'])
        dic_fc['times'].append(timeobj.strftime(dic_cnf['sf_time']))
    logger_wfm.info (str.format(dic_cnf['res_strings']['info']['0002'], len(dic_fc['times'])))
    # Create the rest of the arrays, all but Calls will be empty for now
    for i in range (0, max_days, 1): # restrict to maximum forecast days as calculated before
        # Read date from excel spread sheet and put into array
        row_num = ws.min_row
        col_num = ws.min_column + i + 1
        dateobj = ws.cell(row=row_num, column=col_num).value
        logger_wfm.info(str.format(dic_cnf['res_strings']['info']['0001'], dateobj))
        # Format the dictionary key - Set0, Set1, Set2, ...
        str_tmp = 'Day' + str(i) 
        dic_fc[str_tmp] = {}
        dic_fc[str_tmp]['count'] = max_days
        dic_fc[str_tmp]['date'] = dateobj.strftime(dic_cnf['sf_date'])
        # Insert Call Numbers for this date
        dic_fc[str_tmp]['calls'] = []
        for c in range(0, cnt_Interval):
            r = c + 2
            val = ws.cell(row=r, column=col_num).value
            dic_fc[str_tmp]['calls'].append(int(val))
        dic_fc[str_tmp]['agents'] = []
        dic_fc[str_tmp]['util'] =   []
        dic_fc[str_tmp]['sla'] = []
        dic_fc[str_tmp]['asa'] = []
        dic_fc[str_tmp]['abandon'] = []
        dic_fc[str_tmp]['q-percent'] = []
        dic_fc[str_tmp]['q-time'] = []
        dic_fc[str_tmp]['q-count'] = []

    logger_wfm.info (dic_cnf['res_strings']['info']['0008'])

# ================== First cleanup - effectively from here on only ec and dic_fns is required ==================
    del t1
    del time1
    del str_t1
    del t2
    del str_t2
    del time2
    del timeobj
    del dateobj
    del diff
    del diff_prev
    del str_tmp
    del start_time

# ================== Sanity check of framework data supplied ==================
    logger_wfm.info(dic_cnf['res_strings']['info']['0012'])
    # Check that we are using sound values
    if (dic_cnf['fw']['SLA'] <= 0):
        raise ValueError (str.format(dic_cnf['res_strings']['errors']['0007'], dic_cnf['fw']['SLA'], dic_cnf['f_frame']))
    if (dic_cnf['fw']['AnswerTime'] <= 0):
        raise ValueError (str.format(dic_cnf['res_strings']['errors']['0008'], dic_cnf['fw']['AnswerTime'], dic_cnf['f_frame']))
    if (dic_cnf['fw']['TalkTime'] <= 0):
        raise ValueError (str.format(dic_cnf['res_strings']['errors']['0009'], dic_cnf['fw']['TalkTime'], dic_cnf['f_frame']))
    if (dic_cnf['fw']['ServiceTime'] <= 0):
        raise ValueError (str.format(dic_cnf['res_strings']['errors']['0010'], dic_cnf['fw']['ServiceTime'], dic_cnf['f_frame']))
    if (dic_cnf['fw']['AfterCallWork'] <= 0):
        raise ValueError (str.format(dic_cnf['res_strings']['errors']['0011'], dic_cnf['fw']['AfterCallWork'], dic_cnf['f_frame']))
    if (dic_cnf['fw']['AbandonTime'] <= 0):
        raise ValueError (str.format(dic_cnf['res_strings']['errors']['0012'], dic_cnf['fw']['AbandonTime'], dic_cnf['f_frame']))
    if (dic_cnf['fw']['MaxWait'] <= 0):
        raise ValueError (str.format(dic_cnf['res_strings']['errors']['0013'], dic_cnf['fw']['MaxWait'], dic_cnf['f_frame']))
    
# ================== Create Erlang object and start calculations ================== 
    # (SLA, TTA, ATT, ACW, ABNT, MAX_WAIT, NV, CCC, INTERVAL, OPS_HRS) <- this data comes form the forecast framework file
    ec = Erlang(dic_cnf['fw']['SLA'], dic_cnf['fw']['AnswerTime'], dic_cnf['fw']['TalkTime'], dic_cnf['fw']['AfterCallWork'], dic_cnf['fw']['AbandonTime'], dic_cnf['fw']['MaxWait'], dic_cnf['fw']['NonVoice'], dic_cnf['fw']['Concurrency'], dic_cnf['fw']['ServiceInterval'], dic_cnf['fw']['OperationHours'], avail=dic_cnf['fw']['Availability'])
    max_calls = 0
    for c in range(0,max_days,1):
        str_tmp = 'Day' + str(c)
        i = 0
        for each_call in dic_fc[str_tmp]['calls']:
            # find maximum call volume for Trunk Calculation only if 'summary' flag in conf file is true
            if env['options']['summary'] == True:
                if (each_call > max_calls):
                    max_calls = each_call
                    max_date = dic_fc[str_tmp]['date']
                    max_time = dic_fc['times'][i]
            agents = ec.Agents (dic_cnf['fw']['ServiceTime'], each_call)
            dic_fc[str_tmp]['agents'].append(agents)
            dic_fc[str_tmp]['util'].append(round(ec.Utilisation (dic_cnf['fw']['ServiceTime'], each_call), 2))
            dic_fc[str_tmp]['sla'].append(ec.SLA(agents, each_call, dic_cnf['fw']['ServiceTime']))
            dic_fc[str_tmp]['asa'].append(ec.ASA(agents, each_call))
            dic_fc[str_tmp]['abandon'].append(round(ec.Abandon(agents, each_call), 2))
            dic_fc[str_tmp]['q-percent'].append(round(ec.Queued(agents, each_call), 2))
            dic_fc[str_tmp]['q-time'].append(ec.QueueTime(agents, each_call))
            dic_fc[str_tmp]['q-count'].append(ec.QueueSize(agents, each_call))
            i += 1
    logger_wfm.info(dic_cnf['res_strings']['info']['0014'])
    del each_call
    del str_tmp

# ================== Export the calculated forecast data to JSON ====================
    # Save forcast to JSON file
    if (env['output-format']['json']):
        logger_wfm.info (str.format(dic_cnf['res_strings']['info']['0013'], ['f_result']))
        rec_set = json.dumps(dic_fc)
        with open(dic_cnf['f_result'], 'w') as f:
            json.dump(rec_set, f)
        logger_wfm.info (str.format(dic_cnf['res_strings']['info']['0019'], ['f_result']))
        del rec_set

# ================== Build summary report for max transaction volume =======================
    agents = ec.Agents (dic_cnf['fw']['ServiceTime'], max_calls)
    # Construct Summary Forecast file - only if 'summary' flag is set to 1
    if (env['options']['summary']):
        wfm_helpers.createJSONSummary (agents, max_date, max_time, max_calls, ec, dic_cnf, logger_wfm)
        # Log summary data to log file, if summary-log flag is true    
        if (env['options']['log-summary']):
            wfm_helpers.createLogSummary (agents, max_date, max_time, max_calls, ec, dic_cnf, logger_wfm)
    else:
        logger_wfm.info(dic_cnf['res_strings']['info']['0010'])

# ================== Build Excel Summary Report - Agent count by day by Interval ==================
# TODO: #1 Call other export function here if i.e. datbase export is required
    wfm_helpers.xlCreateReport (env, dic_fc, dic_cnf, logger_wfm)

# ================== Final Cleanup before exit ==================
    del ws
    del wb
    del json_file
    del json
    del os
    del ec
    del dic_fc
    del env
    logger_wfm.info(dic_cnf['res_strings']['info']['0009'])
    del dic_cnf

# ================== Error Handling and Exit ===================================
except ValueError as v_err:
    logging.fatal (v_err, exc_info=True)
except FileNotFoundError as fnf_err:
    logging.fatal (fnf_err, exc_info=True)
except PermissionError as p_err:
    logging.error (p_err, exc_info=True)
except OSError as e:
    logging.error (e, exc_info=True)
except Exception as e:
    logging.fatal (e, exc_info=True)
finally:
    logging.shutdown()

