from pymongo import MongoClient
import logging
from datetime import datetime, timedelta
import wfm_xl
from openpyxl import load_workbook
import uuid

class Wfm_db:
    def __init__ (self, env, str_res, logger):
        try:
            self.str_res = str_res
            self.logger = logger
            self.env    = env
        except Exception as e:
            raise Exception (e)

    def __del__ (self):
        del self.env
        del self.logger

    def dbImportXL (self, dict):
        try:
            self.logger.info (str.format(self.str_res['info']['0029'], self.env['db']['db-name'], self.env['db']['url'], self.env['db']['port']))
            client = MongoClient(self.env['db']['url'], self.env['db']['port'])
            db = client[str(self.env['db']['db-name'])]
            self.logger.info (str.format(self.str_res['info']['0030'], self.env['db']['db-name'], self.env['db']['url'], self.env['db']['port']))
            
            # ================== Create wfm_obj ==================
            xl_obj = wfm_xl.wfm_xl (self.env, dict, self.logger)
            xl_obj.xlPreProcessSource ()
            # Select the active sheet
            ws = xl_obj.getWorksheet()
            collection = db[str(self.env['db']['col-in'])]
            day_cnt = 0
            fld_cnt = 0
            # Process spreadsheet and update db
            # we are moving row by row
            for c in range (2, ws.max_column, 1):
                record = {"_id": '', "date": '', "sdate":"","smonth":"" , "dow":"", "times": [], "calls": [], "util": [], "sla": [], "asa": [], "att": [], "aiw": [], "abandon": [], "qpercent": [], "qtime": [], "qsize": []}
                record['_id']   = str(uuid.uuid1())
                date_cell = ws.cell(1, c).value 
                if (type(date_cell)is datetime and date_cell is not None):
                    record['date']  = date_cell
                    record['sdate'] = datetime.strftime(date_cell, self.env['formats']['dbq-date'])
                    record['dow']   = datetime.strftime(date_cell, '%A')
                    record['smonth'] = datetime.strftime(date_cell, '%b')
                    qry = {"sdate": {"$gt": str(record['sdate'])}}
                    query_res = collection.find_one(qry)
                    if query_res == None:
                        day_cnt += 1
                        for r in range (2, ws.max_row, 1):
                            t1 = ws.cell(r, 1).value
                            record['times'].append (str.format('{}:{}', t1.hour, t1.minute))
                            record['calls'].append (int (round(ws.cell(r, c).value, 0)))
                        if (self.__checkSyntax (record)):
                            collection.insert_one(record)
                            fld_cnt += 1
                    else:
                        self.logger.warning (str.format(dict['res_strings']['warnings']['0002'], datetime.strftime(date_cell, self.env['formats']['date']), db.name, collection.name))
                else:
                    self.logger.warning (str.format(dict['res_strings']['warnings']['0003'], 1, c, ws.cell(r, 1).number_format))
                    break
            self.logger.info (str.format(dict['res_strings']['info']['0034'], day_cnt, fld_cnt, collection.count()))
            del xl_obj
            del record
            del collection
            client.close()
            del client
            del db
            return False
        except ValueError as ve:
            raise ValueError (ve.with_traceback)
        except Exception as e:
            raise Exception (e.with_traceback)

    def __checkSyntax (self, record):
# TODO: add syntax check rules to make sure the data imported makes sense
        return True

